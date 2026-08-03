import os
import re
import traceback
from datetime import date, datetime
from functools import wraps
from io import BytesIO

import pyodbc
from dotenv import load_dotenv
from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash

from usuarios import cargar_usuarios, modificar_usuarios
from clientes import (
    ClientesError,
    cargar_clientes,
    modificar_clientes,
    normalizar_clientes,
)

# Cargar variables de entorno desde .env
load_dotenv()

app = Flask(__name__)

# Clave para firmar las sesiones de Flask (leída desde .env).
app.secret_key = os.getenv("SECRET_KEY", "cambia-esta-clave-en-produccion")

LONGITUD_MIN_PASSWORD = 6
ROLES_VALIDOS = ("admin", "consulta")


# --- Autenticación y autorización ---

def login_required(view):
    """
    Protege una ruta: exige sesión activa. Si el usuario debe cambiar su
    contraseña, lo fuerza a /cambiar-password antes de acceder al módulo.
    """
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"):
            return redirect(url_for("login"))
        if session.get("debe_cambiar"):
            return redirect(url_for("cambiar_password"))
        return view(*args, **kwargs)

    return wrapper


def admin_required(view):
    """Como login_required, pero además exige rol admin (403 si no lo es)."""
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"):
            return redirect(url_for("login"))
        if session.get("debe_cambiar"):
            return redirect(url_for("cambiar_password"))
        if session.get("rol") != "admin":
            return render_template("403.html"), 403
        return view(*args, **kwargs)

    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    """Formulario de acceso. Valida contra usuarios.json."""
    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        password = request.form.get("password") or ""
        usuarios = cargar_usuarios()
        datos = usuarios.get(usuario)

        if datos and check_password_hash(datos.get("password_hash", ""), password):
            session["usuario"] = usuario
            session["rol"] = datos.get("rol", "consulta")
            session["debe_cambiar"] = bool(datos.get("debe_cambiar", False))
            if session["debe_cambiar"]:
                return redirect(url_for("cambiar_password"))
            return redirect(url_for("index"))

        return render_template("login.html", error="Usuario o contraseña incorrectos."), 401

    # Si ya hay sesión, ir directo al inicio.
    if session.get("usuario"):
        return redirect(url_for("index"))
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    """Cierra la sesión."""
    session.clear()
    return redirect(url_for("login"))


@app.route("/cambiar-password", methods=["GET", "POST"])
def cambiar_password():
    """
    Cambio de contraseña. Accesible para cualquier usuario logueado (no se
    bloquea por debe_cambiar, ya que es justamente donde se resuelve).
    """
    if not session.get("usuario"):
        return redirect(url_for("login"))

    if request.method == "POST":
        nueva = request.form.get("password") or ""
        confirmar = request.form.get("confirmar") or ""

        error = None
        if not nueva:
            error = "La contraseña no puede estar vacía."
        elif len(nueva) < LONGITUD_MIN_PASSWORD:
            error = f"La contraseña debe tener al menos {LONGITUD_MIN_PASSWORD} caracteres."
        elif nueva != confirmar:
            error = "Las contraseñas no coinciden."

        if error:
            return render_template("cambiar_password.html", error=error), 400

        usuario = session["usuario"]
        nuevo_hash = generate_password_hash(nueva)

        def _mutar(usuarios):
            if usuario not in usuarios:
                raise ValueError("El usuario ya no existe.")
            usuarios[usuario]["password_hash"] = nuevo_hash
            usuarios[usuario]["debe_cambiar"] = False

        try:
            modificar_usuarios(_mutar)
        except ValueError as exc:
            session.clear()
            return render_template("login.html", error=str(exc)), 400

        session["debe_cambiar"] = False
        return redirect(url_for("index"))

    return render_template("cambiar_password.html", error=None)


# --- Gestión de usuarios (solo admin) ---

@app.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    """Lista usuarios (sin exponer hashes) y ofrece crear/eliminar."""
    return _render_admin()


@app.route("/admin/usuarios/crear", methods=["POST"])
@admin_required
def admin_usuarios_crear():
    """Crea un usuario nuevo con debe_cambiar=True."""
    nombre = (request.form.get("nombre") or "").strip()
    password = request.form.get("password") or ""
    rol = (request.form.get("rol") or "").strip().lower()

    error = None
    if not nombre:
        error = "El nombre de usuario es obligatorio."
    elif len(password) < LONGITUD_MIN_PASSWORD:
        error = f"La contraseña inicial debe tener al menos {LONGITUD_MIN_PASSWORD} caracteres."
    elif rol not in ROLES_VALIDOS:
        error = f"Rol inválido. Use: {', '.join(ROLES_VALIDOS)}."

    if error:
        return _render_admin(error=error), 400

    nuevo_hash = generate_password_hash(password)

    def _mutar(usuarios):
        if nombre in usuarios:
            raise ValueError(f"El usuario '{nombre}' ya existe.")
        usuarios[nombre] = {
            "password_hash": nuevo_hash,
            "rol": rol,
            "debe_cambiar": True,
        }

    try:
        modificar_usuarios(_mutar)
    except ValueError as exc:
        return _render_admin(error=str(exc)), 400

    return _render_admin(mensaje=f"Usuario '{nombre}' creado. Deberá cambiar su contraseña al ingresar.")


@app.route("/admin/usuarios/eliminar", methods=["POST"])
@admin_required
def admin_usuarios_eliminar():
    """Elimina un usuario. No permite que el admin se elimine a sí mismo."""
    nombre = (request.form.get("nombre") or "").strip()

    if nombre == session.get("usuario"):
        return _render_admin(error="No puedes eliminar tu propio usuario."), 400

    def _mutar(usuarios):
        if nombre not in usuarios:
            raise ValueError(f"El usuario '{nombre}' no existe.")
        del usuarios[nombre]

    try:
        modificar_usuarios(_mutar)
    except ValueError as exc:
        return _render_admin(error=str(exc)), 400

    return _render_admin(mensaje=f"Usuario '{nombre}' eliminado.")


def _render_admin(error=None, mensaje=None):
    """Renderiza la vista de gestión de usuarios con la lista actualizada."""
    usuarios = cargar_usuarios()
    lista = sorted(
        (
            {"nombre": nombre, "rol": datos.get("rol", "consulta")}
            for nombre, datos in usuarios.items()
        ),
        key=lambda u: u["nombre"],
    )
    return render_template("admin_usuarios.html", usuarios=lista, error=error, mensaje=mensaje)


# --- Gestión de la relación producto -> cliente(s) (solo admin) ---
# La relación se guarda CIFRADA (clientes.dat). Todas las rutas exigen rol
# admin en el BACKEND (no basta con ocultar el enlace del menú).

@app.route("/admin/clientes")
@admin_required
def admin_clientes():
    """Lista la relación producto->cliente(s) y ofrece agregar/editar/quitar."""
    return _render_admin_clientes()


@app.route("/admin/clientes/guardar", methods=["POST"])
@admin_required
def admin_clientes_guardar():
    """Crea o edita (upsert por COD) una asignación producto->cliente(s)."""
    cod = (request.form.get("cod") or "").strip().upper()
    nombre = (request.form.get("nombre") or "").strip()
    clientes = normalizar_clientes(request.form.get("clientes") or "")

    error = None
    if not cod:
        error = "El código de producto (COD) es obligatorio."
    elif not clientes:
        error = "Debe indicar al menos un cliente (separa varios con + o coma)."
    if error:
        return _render_admin_clientes(error=error), 400

    def _mutar(datos):
        previo = datos.get(cod) or {}
        datos[cod] = {
            "nombre": nombre or (previo.get("nombre") or ""),
            "clientes": clientes,
            # tamano_bolsa se conserva (hoy null); preparado para uso futuro.
            "tamano_bolsa": previo.get("tamano_bolsa"),
        }

    try:
        modificar_clientes(_mutar)
    except ClientesError as exc:
        return _render_admin_clientes(error=str(exc)), 500

    return _render_admin_clientes(mensaje=f"Asignación de '{cod}' guardada.")


@app.route("/admin/clientes/eliminar", methods=["POST"])
@admin_required
def admin_clientes_eliminar():
    """Quita una asignación producto->cliente(s)."""
    cod = (request.form.get("cod") or "").strip().upper()

    def _mutar(datos):
        if cod not in datos:
            raise ValueError(f"La asignación '{cod}' no existe.")
        del datos[cod]

    try:
        modificar_clientes(_mutar)
    except ValueError as exc:
        return _render_admin_clientes(error=str(exc)), 400
    except ClientesError as exc:
        return _render_admin_clientes(error=str(exc)), 500

    return _render_admin_clientes(mensaje=f"Asignación de '{cod}' eliminada.")


def _render_admin_clientes(error=None, mensaje=None):
    """
    Renderiza la gestión de clientes con la relación actualizada. Si la relación
    cifrada no está disponible (clave/archivo), muestra el error y una lista
    vacía sin romper la página.
    """
    lista = []
    error_carga = None
    try:
        datos = cargar_clientes()
        lista = sorted(
            (
                {
                    "cod": cod,
                    "nombre": (registro.get("nombre") or ""),
                    "clientes": (registro.get("clientes") or []),
                }
                for cod, registro in datos.items()
            ),
            key=lambda x: x["cod"],
        )
    except ClientesError as exc:
        error_carga = str(exc)
    return render_template(
        "admin_clientes.html",
        clientes=lista,
        error=error or error_carga,
        mensaje=mensaje,
    )


# --- Configuración de la base de datos (leída desde .env) ---
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "1433")
DB_NAME = os.getenv("DB_NAME", "")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

# Driver recomendado para SQL Server 2012 (v11.0) usando ODBC Driver 17.
# TrustServerCertificate=yes evita errores de validación de certificado.
DB_DRIVER = "ODBC Driver 17 for SQL Server"


def build_connection_string():
    """Construye la cadena de conexión a SQL Server."""
    return (
        f"DRIVER={{{DB_DRIVER}}};"
        f"SERVER={DB_HOST},{DB_PORT};"
        f"DATABASE={DB_NAME};"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
        f"Encrypt=no;"
        f"TrustServerCertificate=yes;"
    )

    # --- ALTERNATIVA si "ODBC Driver 17 for SQL Server" no está instalado ---
    # SQL Server Native Client 11.0 es compatible con SQL Server 2012.
    # Nota: este driver antiguo no admite TrustServerCertificate; usar Encrypt=no.
    # return (
    #     f"DRIVER={{SQL Server Native Client 11.0}};"
    #     f"SERVER={DB_HOST},{DB_PORT};"
    #     f"DATABASE={DB_NAME};"
    #     f"UID={DB_USER};"
    #     f"PWD={DB_PASSWORD};"
    #     f"Encrypt=no;"
    # )


def get_connection():
    """
    Devuelve una conexión reutilizable a SQL Server.

    Lanza pyodbc.Error si la conexión falla (driver ausente, credenciales
    incorrectas, servidor inaccesible, etc.). El llamador es responsable
    de cerrar la conexión.
    """
    try:
        conn = pyodbc.connect(build_connection_string(), timeout=5)
        return conn
    except pyodbc.Error:
        # Re-lanzamos para que el endpoint pueda mostrar el detalle completo.
        raise


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    """
    Verifica la conexión a la base de datos.

    Devuelve 200 si la conexión funciona; 500 con el mensaje de error
    completo de pyodbc si falla (útil para diagnosticar driver o credenciales).
    """
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT @@VERSION;")
        version = cursor.fetchone()[0]
        cursor.close()
        return jsonify({
            "status": "ok",
            "database": DB_NAME,
            "server": f"{DB_HOST},{DB_PORT}",
            "driver": DB_DRIVER,
            "sql_server_version": version,
        }), 200
    except Exception as exc:
        return jsonify({
            "status": "error",
            "database": DB_NAME,
            "server": f"{DB_HOST},{DB_PORT}",
            "driver": DB_DRIVER,
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }), 500
    finally:
        if conn is not None:
            conn.close()


# --- Módulo: Órdenes de Producción ---

@app.route("/ordenes-produccion")
@login_required
def ordenes_produccion():
    """Vista con el formulario de filtro (código de producto + rango de fechas)."""
    return render_template("ordenes_produccion.html")


CABECERA_COLUMNAS = "OP, COD, CANTP, CANTE, CANTM, NOM, CMP, FECHA_I, FECHA_T, NOTA"
DETALLE_COLUMNAS = "OP, COD, CANT, CMP, DOC, LOTE, CANT1, VENCE"

# Nombres legibles para las columnas. Se usan tanto en la vista como en el
# Excel para mantener consistencia. Si una columna no está aquí, se muestra
# su nombre original. Ajusta los alias según la nomenclatura del negocio.
COLUMN_LABELS = {
    "OP": "Orden",
    "COD": "Código",
    "CANTP": "Cant. Programada",
    "CANTE": "Cant. Entregada",
    "CANTM": "Cant. Merma",
    "NOM": "Nombre",
    "CMP": "Componente",
    "FECHA_I": "Fecha Inicio",
    "FECHA_T": "Fecha Término",
    "NOTA": "Nota",
    "CANT": "Cantidad",
    "DOC": "Documento",
    "LOTE": "Lote",
    "CANT1": "Cantidad 1",
    "VENCE": "Vencimiento",
    "OP_OC": "Orden",
}


@app.template_filter("label")
def _label_filter(col):
    """Filtro Jinja: devuelve el alias legible de una columna."""
    return COLUMN_LABELS.get(col, col)


# Prefijos de producto aceptados en el filtro de texto libre (modo producto).
# El usuario escribe uno de estos y la cabecera se consulta como '<prefijo>%'.
# Incluye la película (2A10) y el ducto (2A04) de extrusión PP (desperdicio y
# consumo igual que la manga) y los farmacéuticos 2B/2C (desperdicio por
# movimientos Kardex, ver REGLAS_DESPERDICIO_POR_PRODUCTO).
PREFIJOS_PRODUCTO_VALIDOS = ("2A02", "2A03", "2A04", "2A05", "2A06", "2A10", "2B", "2C")


def _construir_filtro(form):
    """
    Lee y valida el filtro del formulario según el modo elegido y construye la
    condición WHERE parametrizada para la consulta de cabecera.

    Modos:
    - "producto" (por defecto): COD + rango de fechas obligatorio.
      WHERE "COD LIKE ? AND FECHA_I BETWEEN ? AND ?".
    - "orden": uno o varios números de OP. WHERE "OP IN (?, ...)".

    Devuelve (condicion_sql, params, filtro_dict), donde condicion_sql usa
    marcadores '?' (nunca valores) y filtro_dict guarda los valores originales
    para reenviarlos y mostrarlos.

    Lanza ValueError con un mensaje apto para el usuario si la data es inválida.
    """
    modo = (form.get("modo") or "producto").strip()

    if modo == "orden":
        ordenes = []
        for v in form.getlist("ordenes[]"):
            v = (v or "").strip()
            if not v:
                continue
            try:
                ordenes.append(int(v))
            except ValueError:
                raise ValueError("Los números de orden (OP) deben ser numéricos.")
        if not ordenes:
            raise ValueError("Debe indicar al menos un número de orden (OP).")

        marcadores = ", ".join(["?"] * len(ordenes))
        condicion = f"OP IN ({marcadores})"
        filtro = {"modo": "orden", "ordenes": ordenes}
        return condicion, list(ordenes), filtro

    # --- Modo producto (por defecto) ---
    cod = (form.get("cod") or "").strip().upper()
    fecha_desde = (form.get("fecha_desde") or "").strip()
    fecha_hasta = (form.get("fecha_hasta") or "").strip()

    if not cod:
        raise ValueError("Debe indicar el código de producto (COD).")
    # Se acepta cualquier código que EMPIECE por uno de los prefijos permitidos
    # (no exige coincidencia exacta), para poder filtrar por un código completo
    # o una sub-familia (p. ej. "2C01") y no solo por el prefijo corto (COD LIKE
    # '<lo escrito>%' más abajo). Un prefijo no reconocido sigue rechazándose.
    if not any(cod.startswith(p) for p in PREFIJOS_PRODUCTO_VALIDOS):
        raise ValueError("Prefijo no válido. Use 2A02, 2A03, 2A04, 2A05, 2A06, 2A10, 2B o 2C")
    if not fecha_desde or not fecha_hasta:
        raise ValueError("El rango de fechas (desde y hasta) es obligatorio.")

    # El input HTML type=date entrega YYYY-MM-DD. Se convierte a objeto date
    # para pasarlo como parámetro a pyodbc (evita que SQL Server reinterprete
    # el texto y produzca el error 22007 de conversión de fecha).
    try:
        d_desde = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        d_hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("Formato de fecha inválido")

    if d_desde > d_hasta:
        raise ValueError("La fecha 'desde' no puede ser posterior a la fecha 'hasta'.")

    # LIKE por prefijo: se agrega '%' como parámetro, nunca concatenado al SQL.
    # Las fechas van como objetos date, no como string.
    condicion = "COD LIKE ? AND FECHA_I BETWEEN ? AND ?"
    params = [cod + "%", d_desde, d_hasta]
    filtro = {
        "modo": "producto",
        "cod": cod,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    }
    return condicion, params, filtro


def _fetch_dicts(cursor, sql, params):
    """Ejecuta una consulta parametrizada y devuelve filas como lista de dicts."""
    cursor.execute(sql, params)
    columnas = [col[0] for col in cursor.description]
    return [dict(zip(columnas, fila)) for fila in cursor.fetchall()]


# Patrones para extraer "valor + unidad" de un texto tipo "DES X <valor>
# <UND> ...". GENERALIZADO a CUALQUIER unidad (KG, UN, u otras de hasta 6
# letras), no solo KG: la unidad es un grupo opcional/capturado, no literal.
# Se intentan en orden y se usa el primero que matchee:
#   1. "DES X <valor> <UND?> X"  (unidad opcional, con X de cierre) ->
#      "DES X 136.12 KG X ...", "DES X 2488 UN X VARIOS 50743",
#      "DES X 339.7 X" (sin unidad; el grupo opcional queda vacío)
#   2. "DES X <valor> <UND>"     (SIN exigir X de cierre; fallback — la unidad
#      es obligatoria aquí, es lo único que acota el número) ->
#      "DES X 128 KG AJST ARRAN DESC PELUZA F 18"
# Este patrón lo reutilizan tanto la NOTA de la OP (_parse_desperdicio_nota,
# que ignora la unidad: para la nota siempre es kg) como los movimientos
# Kardex de la regla farmacéutica (_sumar_kardex_farmaceutico, que sí necesita
# la unidad real del movimiento, p. ej. "UN").
NOTA_DESPERDICIO_PATRONES = (
    re.compile(r"DES\s*X\s*([\d.,]+)\s*([A-Za-z]{1,6})?\s*X", re.IGNORECASE),
    re.compile(r"DES\s*X\s*([\d.,]+)\s*([A-Za-z]{1,6})\b", re.IGNORECASE),
)


def _parse_desperdicio_valor_unidad(texto):
    """
    Extrae (valor, unidad) de un texto con el patrón "DES X <valor> <UND> ...".
    `unidad` es la que aparece en el texto (p. ej. "KG", "UN"), en mayúsculas,
    o None si el texto no trae unidad (caso "DES X 339.7 X"). Devuelve
    (None, None) si no hay patrón o el valor no es numérico.
    """
    if not texto:
        return None, None
    cadena = str(texto)
    m = None
    for patron in NOTA_DESPERDICIO_PATRONES:
        m = patron.search(cadena)
        if m:
            break
    if not m:
        return None, None
    # El valor puede venir con punto ("23.5") o coma decimal manual ("23,5").
    # Se normaliza a punto sin romper ninguno de los dos formatos. Los valores
    # son < 1000 con punto decimal, así que NO se manejan separadores de miles;
    # si por robustez llegaran ambos separadores, se prioriza el punto como
    # decimal y se descartan las comas (evita que float() falle y devuelva None).
    crudo = m.group(1).strip()
    if "." in crudo:
        crudo = crudo.replace(",", "")
    else:
        crudo = crudo.replace(",", ".")
    try:
        valor = float(crudo)
    except ValueError:
        return None, None
    unidad = (m.group(2) or "").strip().upper() or None
    return valor, unidad


def _parse_desperdicio_nota(nota):
    """
    Extrae el valor de desperdicio (KG) del campo NOTA. Ignora la unidad (para
    la nota de la OP siempre es kg, confirmado con negocio); el patrón está
    generalizado en _parse_desperdicio_valor_unidad, que también reutiliza la
    regla farmacéutica sobre movimientos Kardex (unidad UN u otras).
    Devuelve un float, o None si no se encuentra el patrón o el valor no es
    numérico.
    """
    valor, _unidad = _parse_desperdicio_valor_unidad(nota)
    return valor


def _sumar_cant1_2a(cursor, op):
    """
    Suma CANT1 del detalle (op1) para una OP, filtrando el consumo de códigos
    que empiezan por '2A0' (2A02, 2A03, ... 2A09, incluido 2A08 de mezcla) más
    los '2A62' (excepción, no empieza por 2A0). Devuelve (suma, n_filas). Usa
    parámetros para la OP.

    Este filtro amplio es correcto para TODOS los productos (manga, ducto,
    película, etc.): el detalle de cada orden ya contiene solo lo que esa orden
    realmente consumió, así que 2A0% + 2A62% suma exactamente su consumo.

    UNIDAD: CANT1 de estos códigos está en KILOGRAMOS (kg), la misma unidad que
    el valor de desperdicio de la NOTA (confirmado con negocio). Esta igualdad
    de unidades es la que hace válido el % = kg desperdicio / kg consumo × 100.
    No mezclar con CANTE de la cabecera OP, que está en metros.
    """
    sql = (
        "SELECT OP, COD, CANT, CMP, DOC, LOTE, CANT1 FROM op1 "
        "WHERE OP = ? AND (COD LIKE '2A0%' OR COD LIKE '2A62%')"
    )
    filas = _fetch_dicts(cursor, sql, [op])
    suma = 0
    for f in filas:
        valor = f.get("CANT1")
        if valor is not None:
            suma += valor
    return suma, len(filas)


def _sumar_cant1_patron(cursor, op, patron):
    """
    Suma CANT1 del detalle (op1) de una OP para un ÚNICO patrón LIKE de COD.
    Parametrizada: marcadores para la OP y para el patrón (nada concatenado).
    Devuelve (suma, n_filas). La usan las reglas que necesitan un consumo
    distinto del general (p. ej. el compuesto 2A02 con la mezcla 2A01%).
    """
    sql = "SELECT CANT1 FROM op1 WHERE OP = ? AND COD LIKE ?"
    filas = _fetch_dicts(cursor, sql, [op, patron])
    suma = 0
    for f in filas:
        valor = f.get("CANT1")
        if valor is not None:
            suma += valor
    return suma, len(filas)


# --- Kardex (regla farmacéutica, prefijos 2B/2C) ---------------------------
# Para estos productos la NOTA de la OP está bloqueada; el desperdicio se lee
# de los movimientos de inventario Kardex, en tres tablas con los MISMOS
# campos (COD, CANT, OP_OC, DES) separadas por ANTIGÜEDAD de la OP. Antes se
# consultaban las tres con UNION ALL por prudencia (no se sabía en cuál caía
# cada orden); validado con datos reales que la partición depende de la
# FECHA_T (término) de la propia OP, así que ahora se resuelve UNA sola tabla
# por orden (ver _tabla_kardex_por_fecha) — Kardexhi tiene ~4M filas y
# escanearla siempre añadía minutos por OP. Solo lectura.
KARDEX_TABLAS = ("Kardex", "KardexA", "Kardexhi")
KARDEX_TABLAS_VALIDAS = set(KARDEX_TABLAS)


def _tabla_kardex_por_fecha(fecha_t, hoy=None):
    """
    Resuelve en cuál de las tres tablas Kardex está una OP farmacéutica, según
    su FECHA_T (fecha de TÉRMINO; FECHA_I NO sirve para esto), relativa a la
    fecha actual — las tres particionan por antigüedad (validado con datos
    reales: OP 461221 FECHA_T=2026-07-06 -> "Kardex"; OP 445393
    FECHA_T=2025-12-22 -> "Kardexhi"):

        mismo año y mismo mes que hoy   -> "Kardex"
        mismo año, mes ANTERIOR a hoy   -> "KardexA"
        año ANTERIOR a hoy              -> "Kardexhi"

    Se compara contra la fecha actual en cada ejecución (nunca un valor
    fijo). Devuelve None si fecha_t es nula o si no encaja en ninguna de las
    tres ventanas (p. ej. una fecha futura, que no debería darse en una OP ya
    terminada) — la orden queda "sin cálculo" con alerta, sin reventar.
    `hoy` es inyectable para pruebas; por defecto date.today().
    """
    if fecha_t is None:
        return None
    if hoy is None:
        hoy = date.today()
    if fecha_t.year == hoy.year and fecha_t.month == hoy.month:
        return "Kardex"
    if fecha_t.year == hoy.year and fecha_t.month < hoy.month:
        return "KardexA"
    if fecha_t.year < hoy.year:
        return "Kardexhi"
    return None

# Detecta si un movimiento Kardex es una fila de desperdicio: su campo DES,
# ya normalizado con strip() (tolera espacios al inicio/fin del dato crudo),
# EMPIEZA con "DES X" (ancla ^ sobre el texto normalizado — a diferencia de
# NOTA_DESPERDICIO_PATRONES que busca en cualquier posición, aquí NUNCA se
# detecta "DES X" en mitad de otro texto). No se filtra por COD: la selección
# es por el contenido de DES, la orden puede tener filas de todo tipo en
# Kardex.
FILA_DESPERDICIO_PATRON = re.compile(r"^DES\s*X", re.IGNORECASE)


def _es_fila_desperdicio(des):
    """True si el campo DES (normalizado con strip()) de un movimiento Kardex
    es una fila de desperdicio."""
    return bool(des) and bool(FILA_DESPERDICIO_PATRON.match(str(des).strip()))


KARDEX_COLUMNAS = "OP_OC, COD, CANT, VMP, DOC, LOTE, DES"


def _consultar_kardex_op(cursor, op, tabla):
    """
    Trae los movimientos Kardex de una orden por OP_OC, de UNA sola tabla (la
    resuelta por _tabla_kardex_por_fecha según la FECHA_T de la OP) en vez del
    UNION ALL sobre las tres — Kardexhi tiene ~4M filas y escanearla siempre
    para las tres tablas tardaba minutos por OP.

    `tabla` DEBE venir del resultado de _tabla_kardex_por_fecha (nunca de
    entrada de usuario). El nombre de tabla no puede ir parametrizado con '?'
    en T-SQL, así que se valida contra KARDEX_TABLAS_VALIDAS (un set fijo) y
    solo entonces se interpola — la OP_OC sí va parametrizada. Solo lectura.
    """
    if tabla not in KARDEX_TABLAS_VALIDAS:
        raise ValueError(f"Tabla Kardex no válida: {tabla!r}")
    sql = f"SELECT {KARDEX_COLUMNAS} FROM {tabla} WHERE OP_OC = ?"
    return _fetch_dicts(cursor, sql, [op])


def _sumar_kardex_farmaceutico(cursor, op, fecha_t):
    """
    Numerador/denominador de la regla farmacéutica para una orden, a partir de
    sus movimientos Kardex de UNA sola tabla (ver _tabla_kardex_por_fecha,
    resuelta según la FECHA_T de la propia OP). Solo cuentan las filas cuyo
    DES empieza con "DES X" (_es_fila_desperdicio); de esas, se parsea el
    valor y la unidad con _parse_desperdicio_valor_unidad. Si una fila DES no
    parsea, se IGNORA por completo (ni su valor ni su CANT cuentan):

        num = Σ(valor parseado de DES)     de las filas que sí parsearon
        den = Σ(CANT)                      de esas mismas filas

    Devuelve (num, den, n_filas_des, n_filas_usadas, unidad, filas, tabla):
    - n_filas_des: total de filas con patrón "DES X" encontradas (diagnóstico).
    - n_filas_usadas: cuántas de esas efectivamente parsearon (num/den las
      incluyen; n_filas_des - n_filas_usadas se ignoraron).
    - unidad: unidad común de las filas usadas (KG, UN, ...), o None si no hay
      ninguna o si las filas usadas mezclan unidades distintas (ambiguo).
    - filas: SOLO las filas Kardex de la OP con patrón "DES X" (no el resto de
      movimientos), para mostrar en el detalle de la vista — cambio
      puramente visual, el cálculo de num/den es el mismo con o sin este
      filtro. Cada dict trae una clave extra "_usada" (bool): True si esa
      fila efectivamente PARSEÓ y entró en num/den, False si tiene el patrón
      "DES X" pero no se pudo parsear. No es una columna real de Kardex.
    - tabla: la tabla Kardex efectivamente consultada, o None si no se pudo
      resolver (FECHA_T nula o fuera de las tres ventanas) — en ese caso no se
      ejecuta ninguna consulta y se devuelve todo vacío, sin reventar.
    """
    tabla = _tabla_kardex_por_fecha(fecha_t)
    if tabla is None:
        return 0.0, 0.0, 0, 0, None, [], None

    filas = _consultar_kardex_op(cursor, op, tabla)

    num = 0.0
    den = 0.0
    n_filas_des = 0
    n_usadas = 0
    unidades = set()
    filas_des = []
    for f in filas:
        if not _es_fila_desperdicio(f.get("DES")):
            continue  # no es fila de desperdicio: ni se cuenta ni se muestra
        n_filas_des += 1
        f["_usada"] = False
        valor, unidad = _parse_desperdicio_valor_unidad(f.get("DES"))
        if valor is not None:
            cant = f.get("CANT")
            num += valor
            den += float(cant) if cant is not None else 0.0
            n_usadas += 1
            f["_usada"] = True
            if unidad:
                unidades.add(unidad)
        filas_des.append(f)

    unidad_comun = unidades.pop() if len(unidades) == 1 else None
    return num, den, n_filas_des, n_usadas, unidad_comun, filas_des, tabla


def _calcular_desperdicio(nota, suma_cant1, n_filas_2a):
    """
    Regla GENERAL de cálculo del porcentaje de desperdicio:
        porcentaje = valor_nota / suma_consumo * 100

    - valor_nota (KG): se extrae de la NOTA.
    - suma_cant1: suma de CANT1 del detalle op1 con COD LIKE '2A0%' o '2A62%'
      (base=100%).

    UNIDADES: valor_nota (kg) y suma_cant1 (kg) están en la MISMA unidad
    (kilogramos), confirmado con negocio; por eso el cociente da un % válido.

    suma_cant1 llega como decimal.Decimal desde pyodbc; se convierte a float
    antes de dividir. Devuelve un dict con los datos para la vista, incluyendo
    una alerta cuando no es posible calcular.

    Esta es la regla por defecto; el punto de entrada por orden es
    _calcular_desperdicio_producto, que selecciona la regla según el producto.
    """
    valor_nota = _parse_desperdicio_nota(nota)
    suma_f = float(suma_cant1) if suma_cant1 is not None else 0.0

    resultado = {
        "valor_nota": valor_nota,
        "suma_cant1": suma_cant1,
        "porcentaje": None,
        "alerta": None,
    }
    if valor_nota is None:
        resultado["alerta"] = "Formato de nota no encontrado"
        return resultado
    if n_filas_2a == 0 or suma_f <= 0:
        resultado["alerta"] = "Sin consumo (2A0/2A62) para calcular"
        return resultado

    resultado["porcentaje"] = round((float(valor_nota) / suma_f) * 100, 2)
    return resultado


# --- Reglas de desperdicio por producto ------------------------------------
# La mayoría de productos usan la regla GENERAL (leen el desperdicio de la
# NOTA). Algunos lo calculan distinto y se registran por prefijo de COD
# (startswith). Todas las reglas comparten una MISMA firma basada en un
# contexto (ctx) para poder intercambiarse desde el punto único de cálculo:
#     regla(ctx) -> dict(valor_nota, suma_cant1, porcentaje, alerta, metodo,
#                        desperdicio_kg, unidad, num_ponderado, den_ponderado)
# ctx trae todo lo que cualquier regla pueda necesitar (nota, consumo
# 2A0%+2A62%, CANTE de la cabecera, consumo de mezcla 2A01% y movimientos
# Kardex para la regla farmacéutica).
#
# Campos canónicos que TODA regla puebla para el resto de la app:
#   - metodo: 'nota' (leído de la nota) | 'diferencia' (consumido-entregado,
#     compuesto 2A02) | 'kardex' (leído de movimientos Kardex, farmacéutico
#     2B/2C). Sirve para mostrar etiquetas correctas.
#   - desperdicio_kg: kg (o unidad equivalente, ver `unidad`) de desperdicio
#     de la orden (None si no calculable).
#   - unidad: unidad del desperdicio/consumo de la orden ('KG' para nota y
#     diferencia, confirmado con negocio; la unidad real del movimiento para
#     kardex — 'UN', 'KG', etc. — o None si no se pudo determinar). Los %
#     siempre son comparables entre unidades; los kg/unidades ABSOLUTOS NO
#     deben sumarse entre órdenes de distinta unidad.
#   - num_ponderado / den_ponderado: numerador y denominador con que la orden
#     entra al PONDERADO de su grupo (None si no debe entrar). Regla por nota:
#     num = kg de la nota, den = consumo. Regla 2A02: num = kg por diferencia
#     (>=0), den = entregado (CANTE). Regla farmacéutica: num = Σ DES
#     parseado, den = Σ CANT de esas filas. Así Σnum/Σden×100 es coherente
#     entre métodos aunque un grupo (p. ej. por mes) mezcle varios de ellos.


def _regla_nota(ctx):
    """
    Regla GENERAL (manga 2A03, ducto 2A04, película 2A10 y demás): el
    desperdicio se LEE de la nota. Adapta _calcular_desperdicio (que NO se
    toca) al contrato de contexto y agrega los campos canónicos.
    """
    res = _calcular_desperdicio(ctx.get("nota"), ctx.get("suma_cant1"),
                                ctx.get("n_filas_2a"))
    res["metodo"] = "nota"
    res["desperdicio_kg"] = res.get("valor_nota")
    res["unidad"] = "KG"
    # Entra al ponderado con num = kg de la nota, den = consumo (2A0%+2A62%).
    # El gate den > 0 lo aplica cada punto de agregación (excluye consumo 0 y
    # órdenes sin nota, igual que antes).
    res["num_ponderado"] = res.get("valor_nota")
    res["den_ponderado"] = res.get("suma_cant1")
    return res


def _calcular_desperdicio_2a02(entregado, suma_2a01, n_filas_2a01):
    """
    Regla del COMPUESTO de mezcla/peletizado (prefijo 2A02). NO lee la nota:
    el desperdicio se obtiene por DIFERENCIA entre consumo y producción.

        entregado (producido) = CANTE de la cabecera OP (kg).
        consumido             = Σ CANT1 del detalle op1 con COD LIKE '2A01%'
                                (mezcla; NO el 2A0%+2A62% de la regla general).
        desperdicio (kg)      = consumido - entregado.
        % = desperdicio / entregado * 100      (denominador = ENTREGADO).

    Validación de referencia (OP 436615): entregado 3570, consumido 3672.40 ->
    desperdicio 102.40 -> 102.40/3570*100 = 2.87% (NO 2.79%, que sería sobre lo
    consumido; el denominador es lo ENTREGADO).
    """
    consumido = float(suma_2a01) if suma_2a01 is not None else 0.0
    ent = float(entregado) if entregado is not None else 0.0

    resultado = {
        "valor_nota": None,          # esta regla NO usa la nota
        "suma_cant1": suma_2a01,     # consumo relevante = mezcla 2A01
        "porcentaje": None,
        "alerta": None,
        "metodo": "diferencia",
        "entregado": entregado,      # CANTE (kg producido)
        "desperdicio_kg": None,
        "unidad": "KG",
        "num_ponderado": None,
        "den_ponderado": None,
    }
    # Orden 2A02 sin filas 2A01 (consumido = 0): no forzar un número.
    if n_filas_2a01 == 0 or consumido <= 0:
        resultado["alerta"] = "Sin consumo de mezcla (2A01) para calcular"
        return resultado
    # Entregado (CANTE) = 0: no dividir (mismo criterio que la división por
    # cero de la regla general).
    if ent <= 0:
        resultado["alerta"] = "Sin cantidad entregada (CANTE) para calcular"
        return resultado

    desperdicio = consumido - ent
    # Consumido < entregado -> desperdicio negativo, sin sentido físico: se
    # reporta 0% (no se muestran porcentajes negativos) y la orden queda
    # visible. Decisión de negocio confirmada.
    if desperdicio < 0:
        desperdicio = 0.0

    resultado["desperdicio_kg"] = round(desperdicio, 2)
    resultado["porcentaje"] = round((desperdicio / ent) * 100, 2)
    # Ponderado: num = kg de desperdicio (>=0), den = entregado (CANTE).
    resultado["num_ponderado"] = desperdicio
    resultado["den_ponderado"] = ent
    return resultado


def _regla_compuesto_2a01(ctx):
    """Adapta la regla del compuesto (2A02) al contrato de contexto."""
    return _calcular_desperdicio_2a02(ctx.get("cante"), ctx.get("suma_2a01"),
                                      ctx.get("n_filas_2a01"))


def _calcular_desperdicio_farmaceutico(num_kardex, den_kardex, n_filas_des,
                                       n_filas_usadas, unidad, tabla_kardex=None):
    """
    Regla FARMACÉUTICA (prefijos 2B/2C). La NOTA de la OP está BLOQUEADA para
    estos productos; el desperdicio se lee de los movimientos Kardex (ver
    _sumar_kardex_farmaceutico, hoy UNA sola tabla resuelta por FECHA_T, ver
    _tabla_kardex_por_fecha), en las filas cuyo DES tiene el patrón
    "DES X <valor> <UND> ...":

        % = Σ(DES parseado) / Σ(CANT) × 100

    UNIDADES: DES y CANT de una misma fila están en la MISMA unidad
    (confirmado), por eso el % es válido; pero productos distintos pueden
    tener unidades distintas (UN, KG, ...) — el % es lo único agregable entre
    órdenes; el desperdicio absoluto NUNCA debe sumarse entre unidades.

    Validación de referencia (OP_OC 463021): DES "DES X 2488 UN X VARIOS
    50743" -> valor 2488, unidad UN.
    """
    resultado = {
        "valor_nota": None,          # esta regla NO usa la nota (bloqueada)
        "suma_cant1": den_kardex,    # denominador = Σ CANT de las filas DES
        "porcentaje": None,
        "alerta": None,
        "metodo": "kardex",
        "desperdicio_kg": None,      # canónico; aquí puede NO ser kg (ver unidad)
        "unidad": unidad,
        "n_filas_des": n_filas_des,
        "num_ponderado": None,
        "den_ponderado": None,
        "tabla_kardex": tabla_kardex,  # diagnóstico: tabla efectivamente usada
    }
    # Sin tabla resoluble (FECHA_T nula o fuera de las tres ventanas): no se
    # llegó a consultar Kardex, se distingue del caso "sin filas DES".
    if tabla_kardex is None:
        resultado["alerta"] = "Sin FECHA_T de la orden: no se pudo determinar la tabla Kardex"
        return resultado
    # Sin filas "DES X" en Kardex, o ninguna parseable: no forzar un número.
    if n_filas_usadas == 0:
        resultado["alerta"] = "Sin filas de desperdicio (DES X) en Kardex para calcular"
        return resultado
    den_f = float(den_kardex) if den_kardex is not None else 0.0
    if den_f <= 0:
        resultado["alerta"] = "Sin consumo (Σ CANT) en Kardex para calcular"
        return resultado

    resultado["desperdicio_kg"] = round(float(num_kardex), 2)
    resultado["porcentaje"] = round((float(num_kardex) / den_f) * 100, 2)
    resultado["num_ponderado"] = float(num_kardex)
    resultado["den_ponderado"] = den_f
    return resultado


def _regla_farmaceutica(ctx):
    """Adapta la regla farmacéutica (Kardex, 2B/2C) al contrato de contexto."""
    return _calcular_desperdicio_farmaceutico(
        ctx.get("num_kardex"), ctx.get("den_kardex"),
        ctx.get("n_filas_des", 0), ctx.get("n_filas_kardex_usadas", 0),
        ctx.get("unidad_kardex"), ctx.get("tabla_kardex"))


# Registro de reglas por prefijo de producto (COD). Coincidencia por startswith
# (la primera que coincide gana). Prefijo sin regla -> regla general por nota.
REGLAS_DESPERDICIO_POR_PRODUCTO = {
    "2A02": _regla_compuesto_2a01,   # compuesto mezcla/peletizado: por diferencia
    "2B": _regla_farmaceutica,       # farmacéutico: por movimientos Kardex
    "2C": _regla_farmaceutica,       # farmacéutico: por movimientos Kardex
}


def _regla_desperdicio(cod):
    """
    Devuelve la regla (función que recibe ctx) aplicable a un producto según su
    COD. Si el prefijo está en REGLAS_DESPERDICIO_POR_PRODUCTO usa esa; si no,
    la regla general por nota (_regla_nota).
    """
    if cod:
        cod_txt = str(cod).strip().upper()
        for prefijo, regla in REGLAS_DESPERDICIO_POR_PRODUCTO.items():
            if cod_txt.startswith(prefijo):
                return regla
    return _regla_nota


def _calcular_desperdicio_producto(cod, nota, suma_cant1, n_filas_2a,
                                   cante=None, suma_2a01=None, n_filas_2a01=0,
                                   num_kardex=None, den_kardex=None,
                                   n_filas_des=0, n_filas_kardex_usadas=0,
                                   unidad_kardex=None, tabla_kardex=None):
    """
    Punto ÚNICO de cálculo del desperdicio por orden. Arma el contexto con todo
    lo que cualquier regla pueda necesitar y delega en la regla del producto
    (selección por prefijo). El compuesto (2A02) calcula por diferencia
    (CANTE vs. consumo de mezcla 2A01); el farmacéutico (2B/2C) por movimientos
    Kardex; el resto lee la nota.
    """
    ctx = {
        "cod": cod,
        "nota": nota,
        "suma_cant1": suma_cant1,
        "n_filas_2a": n_filas_2a,
        "cante": cante,
        "suma_2a01": suma_2a01,
        "n_filas_2a01": n_filas_2a01,
        "num_kardex": num_kardex,
        "den_kardex": den_kardex,
        "n_filas_des": n_filas_des,
        "n_filas_kardex_usadas": n_filas_kardex_usadas,
        "unidad_kardex": unidad_kardex,
        "tabla_kardex": tabla_kardex,
    }
    regla = _regla_desperdicio(cod)
    return regla(ctx)


def _consultar_ordenes(form):
    """
    Ejecuta las consultas de cabecera (según el modo del filtro) y detalle, y
    devuelve los resultados agrupados por número de orden (OP).

    Usa siempre consultas parametrizadas. Lanza ValueError si el filtro es
    inválido y pyodbc.Error si falla la base de datos.
    """
    condicion, params_cabecera, _ = _construir_filtro(form)

    sql_cabecera = f"SELECT {CABECERA_COLUMNAS} FROM OP WHERE {condicion}"

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cabecera_filas = _fetch_dicts(cursor, sql_cabecera, params_cabecera)

        # El detalle (op1) se trae para las órdenes encontradas en la cabecera.
        ops = [fila["OP"] for fila in cabecera_filas]
        if ops:
            marcadores = ", ".join(["?"] * len(ops))
            sql_detalle = (
                f"SELECT {DETALLE_COLUMNAS} FROM op1 "
                f"WHERE OP IN ({marcadores}) AND (COD LIKE '2A0%' OR COD LIKE '2A62%')"
            )
            detalle_filas = _fetch_dicts(cursor, sql_detalle, ops)
        else:
            detalle_filas = []

        # Suma de CANT1 (consumo 2A0%+2A62%) por cada OP, para el desperdicio.
        sumas_2a = {op: _sumar_cant1_2a(cursor, op) for op in ops}

        # Consumo de mezcla (2A01%) SOLO para las órdenes cuya regla lo requiere
        # (hoy el compuesto 2A02, que calcula por diferencia). Se detecta por la
        # misma regla registrada, evitando la consulta extra para los demás.
        cod_por_op = {fila["OP"]: fila.get("COD") for fila in cabecera_filas}
        sumas_2a01 = {
            op: _sumar_cant1_patron(cursor, op, "2A01%")
            for op in ops
            if _regla_desperdicio(cod_por_op.get(op)) is _regla_compuesto_2a01
        }

        # Movimientos Kardex SOLO para las órdenes cuya regla lo requiere (hoy
        # el farmacéutico 2B/2C, que calcula sobre las filas "DES X" de UNA
        # sola tabla Kardex/KardexA/Kardexhi, resuelta por la FECHA_T de cada
        # OP — ver _tabla_kardex_por_fecha). Se detecta por la misma regla
        # registrada, evitando la consulta extra (de solo lectura) para los
        # demás.
        fecha_t_por_op = {fila["OP"]: fila.get("FECHA_T") for fila in cabecera_filas}
        sumas_kardex = {
            op: _sumar_kardex_farmaceutico(cursor, op, fecha_t_por_op.get(op))
            for op in ops
            if _regla_desperdicio(cod_por_op.get(op)) is _regla_farmaceutica
        }
        cursor.close()
    finally:
        if conn is not None:
            conn.close()

    # Agrupar cabecera y detalle por número de orden (OP).
    agrupado = {}

    def _bucket(op):
        return agrupado.setdefault(
            op, {"op": op, "cabecera": None, "detalle": [], "detalle_kardex": []})

    for fila in cabecera_filas:
        _bucket(fila["OP"])["cabecera"] = fila
    for fila in detalle_filas:
        _bucket(fila["OP"])["detalle"].append(fila)
    for op, tupla_kardex in sumas_kardex.items():
        _bucket(op)["detalle_kardex"] = tupla_kardex[5]  # posición 5: filas

    # Calcular el desperdicio de cada orden (regla según el producto).
    for op, orden in agrupado.items():
        suma, n_filas = sumas_2a.get(op, (0, 0))
        suma01, n01 = sumas_2a01.get(op, (0, 0))
        num_k, den_k, n_des, n_usadas, unidad_k, _filas_k, tabla_k = sumas_kardex.get(
            op, (0.0, 0.0, 0, 0, None, [], None))
        cab = orden["cabecera"]
        nota = cab.get("NOTA") if cab else None
        cod = cab.get("COD") if cab else None
        cante = cab.get("CANTE") if cab else None
        orden["desperdicio"] = _calcular_desperdicio_producto(
            cod, nota, suma, n_filas,
            cante=cante, suma_2a01=suma01, n_filas_2a01=n01,
            num_kardex=num_k, den_kardex=den_k, n_filas_des=n_des,
            n_filas_kardex_usadas=n_usadas, unidad_kardex=unidad_k,
            tabla_kardex=tabla_k)

    # Ordenar por número de OP para una salida estable.
    return [agrupado[k] for k in sorted(agrupado)]


def _extraer_filtro(form):
    """
    Extrae el filtro en un dict simple para reenviarlo (campos hidden) desde
    la vista de resultados hacia la ruta de exportación.
    """
    _, _, filtro = _construir_filtro(form)
    return filtro


def _porcentaje_ponderado(num, den):
    """
    % de desperdicio ponderado de un grupo: suma(valor_nota) / suma(CANT1) * 100.
    Devuelve None si no hay base (den <= 0), coherente con _calcular_desperdicio.
    """
    if den and den > 0:
        return round((num / den) * 100, 2)
    return None


def _agrupar_ponderado(ordenes, clave):
    """
    Agrupa las órdenes ya consultadas (en memoria, sin tocar SQL) según la
    función `clave(orden) -> (id_grupo, etiqueta)` y calcula por grupo el
    desperdicio PONDERADO: suma(valor_nota) / suma(suma_cant1) * 100, NO el
    promedio de los % individuales.

    Solo acumula en el ponderado las órdenes con valor_nota y suma_cant1
    válidos (las mismas que producen un porcentaje en _calcular_desperdicio);
    `n_ordenes` cuenta todas las órdenes del grupo. Si `clave` devuelve None
    para una orden, esa orden se omite del agrupado.

    Devuelve una lista de dicts: {id, etiqueta, n_ordenes, porcentaje}.
    """
    grupos = {}
    for orden in ordenes:
        info = clave(orden)
        if info is None:
            continue
        gid, etiqueta = info
        g = grupos.get(gid)
        if g is None:
            g = grupos[gid] = {
                "id": gid,
                "etiqueta": etiqueta,
                "n_ordenes": 0,
                "num": 0.0,
                "den": 0.0,
            }
        g["n_ordenes"] += 1

        desp = orden.get("desperdicio") or {}
        num = desp.get("num_ponderado")
        den = desp.get("den_ponderado")
        # DEFINICIÓN DEL PONDERADO: cada orden aporta un numerador (kg de
        # desperdicio) y un denominador (su base) que fija cada regla:
        #   - Regla por nota: num = kg de la nota, den = consumo (2A0%+2A62%).
        #   - Regla 2A02 (por diferencia): num = kg (consumido-entregado, >=0),
        #     den = entregado (CANTE).
        # Solo entran las órdenes con base den > 0 (las mismas que producen un %
        # por orden); las órdenes sin nota, sin consumo o sin entregado NO se
        # incluyen (ni su num ni su den). Un grupo por mes/cliente puede mezclar
        # ambos métodos: el ponderado sigue siendo Σnum/Σden×100 y es coherente
        # porque en ambos num=kg de desperdicio y den=base de la orden.
        if num is not None and den is not None and float(den) > 0:
            g["num"] += float(num)
            g["den"] += float(den)

    filas = []
    for g in grupos.values():
        filas.append({
            "id": g["id"],
            "etiqueta": g["etiqueta"],
            "n_ordenes": g["n_ordenes"],
            "porcentaje": _porcentaje_ponderado(g["num"], g["den"]),
        })
    return filas


def _clave_producto(orden):
    """Clave de agrupación por producto: usa COD/NOM de la cabecera."""
    cab = orden.get("cabecera") or {}
    cod = cab.get("COD")
    nom = cab.get("NOM")
    cod_txt = str(cod).strip() if cod is not None else ""
    nom_txt = str(nom).strip() if nom is not None else ""
    etiqueta = " — ".join(p for p in (cod_txt, nom_txt) if p) or "Sin producto"
    return (cod_txt or "—", etiqueta)


def _hacer_clave_cliente(mapa_clientes):
    """
    Devuelve una función clave(orden) -> (id_grupo, etiqueta) para agrupar
    "Por cliente" (Opción B: grupo conjunto). El grupo se determina por el
    producto (COD) de la orden vía la relación producto->cliente(s):

    - Producto con varios clientes (p. ej. ["FMCA","BEKER"]) -> un ÚNICO grupo
      "FMCA + BEKER" (clientes unidos por " + ", en el orden almacenado). Las
      referencias compartidas NO se suman a cada cliente por separado.
    - Producto con un cliente -> grupo con ese nombre.
    - Producto que NO está en la relación -> "Sin cliente asignado".

    Cada orden cae en EXACTAMENTE un grupo, así la suma de los grupos iguala el
    total (sin doble conteo). El % lo calcula _agrupar_ponderado igual que el
    resto (Σ num / Σ den × 100, respetando nota/diferencia).
    """
    def _clave_cliente(orden):
        cab = orden.get("cabecera") or {}
        cod = cab.get("COD")
        cod_txt = str(cod).strip() if cod is not None else ""
        registro = mapa_clientes.get(cod_txt) if cod_txt else None
        clientes = registro.get("clientes") if isinstance(registro, dict) else None
        if clientes:
            etiqueta = " + ".join(clientes)
            return (etiqueta, etiqueta)
        return ("Sin cliente asignado", "Sin cliente asignado")

    return _clave_cliente


def _clave_mes(orden):
    """Clave de agrupación por año-mes de FECHA_I."""
    cab = orden.get("cabecera") or {}
    fecha = cab.get("FECHA_I")
    if isinstance(fecha, datetime):
        gid = f"{fecha.year:04d}-{fecha.month:02d}"
        return (gid, gid)
    if hasattr(fecha, "year") and hasattr(fecha, "month"):
        gid = f"{fecha.year:04d}-{fecha.month:02d}"
        return (gid, gid)
    # Sin fecha utilizable: se agrupa aparte y se ordena al final.
    return ("9999-99", "Sin fecha")


def _clave_dia(orden):
    """Clave de agrupación por día (YYYY-MM-DD) de FECHA_I."""
    cab = orden.get("cabecera") or {}
    fecha = cab.get("FECHA_I")
    if isinstance(fecha, datetime) or (
        hasattr(fecha, "year") and hasattr(fecha, "month") and hasattr(fecha, "day")
    ):
        gid = f"{fecha.year:04d}-{fecha.month:02d}-{fecha.day:02d}"
        return (gid, gid)
    # Sin fecha utilizable: se agrupa aparte y se ordena al final.
    return ("9999-99-99", "Sin fecha")


def _fechas_validas(ordenes):
    """
    Devuelve las FECHA_I utilizables como objetos date (normaliza datetime a
    date para poder medir el rango sin mezclar tipos).
    """
    fechas = []
    for orden in ordenes:
        cab = orden.get("cabecera") or {}
        f = cab.get("FECHA_I")
        if isinstance(f, datetime):
            fechas.append(f.date())
        elif hasattr(f, "year") and hasattr(f, "month") and hasattr(f, "day"):
            fechas.append(f)
    return fechas


# Umbral (en días) para decidir la granularidad del gráfico temporal: si el
# rango de la consulta supera este valor se agrupa por mes; si no, por día.
UMBRAL_DIAS_AGRUPAR_MES = 60


def _tendencia_por_grupo(ordenes, orden_productos, agrupacion):
    """
    Serie temporal por producto para el gráfico de líneas del tab Resumen:
    una línea por producto, con el % de desperdicio ponderado en cada periodo.

    Reutiliza _agrupar_ponderado con una clave compuesta (producto, periodo);
    NO lanza consultas nuevas a SQL Server. `orden_productos` indica el orden
    en que deben salir los productos (el mismo del gráfico de barras) para que
    los colores de la paleta sean coherentes entre ambos gráficos.

    Devuelve {labels, series, agrupacion, suficiente}. `suficiente` es False si
    no hay al menos dos periodos que graficar (una línea necesita >= 2 puntos).
    """
    clave_periodo = _clave_dia if agrupacion == "día" else _clave_mes

    def _clave(orden):
        pid, plabel = _clave_producto(orden)
        perid, perlabel = clave_periodo(orden)
        return ((pid, perid), (plabel, perlabel))

    celdas = _agrupar_ponderado(ordenes, _clave)

    # Periodos del eje X, ordenados cronológicamente por su id.
    periodos = {}
    valores = {}
    for c in celdas:
        _, perid = c["id"]
        plabel, perlabel = c["etiqueta"]
        periodos.setdefault(perid, perlabel)
        valores[(plabel, perid)] = c["porcentaje"]

    periodos_orden = sorted(periodos)
    labels = [periodos[p] for p in periodos_orden]

    # Una serie por producto, respetando el orden recibido y solo con productos
    # que tengan al menos un punto con % calculable (evita líneas vacías).
    series = []
    for plabel in orden_productos:
        data = [valores.get((plabel, perid)) for perid in periodos_orden]
        if any(v is not None for v in data):
            series.append({"producto": plabel, "data": data})

    return {
        "labels": labels,
        "series": series,
        "agrupacion": agrupacion,
        "suficiente": len(labels) >= 2 and len(series) >= 1,
    }


# Regla de color del % de desperdicio (única fuente de verdad en backend):
#   < 10%  -> verde (aceptable) · 10–15% -> amber (atención) · > 15% -> naranja
# El macro Jinja `nivel_pct` replica esta regla para el HTML puro.
def _nivel_desperdicio(pct):
    """Devuelve 'verde' | 'amber' | 'naranja' según el % (None -> None)."""
    if pct is None:
        return None
    if pct < 10:
        return "verde"
    if pct <= 15:
        return "amber"
    return "naranja"


# Colores de traza para las sparklines (coherentes con los tokens CSS).
_NIVEL_HEX = {"verde": "#0E9F6E", "amber": "#E8A317", "naranja": "#F07A1E"}
_GRIS_SPARK = "#C7CAD1"
_GRIS_SPARK_DOT = "#9296A1"

# Umbral de pendiente (|slope| en % por orden) para considerar la tendencia
# "estable" en lugar de "sube"/"baja".
UMBRAL_PENDIENTE = 0.1


def _sparkline(puntos, nivel):
    """
    Construye una mini-línea (sparkline) SVG a partir de la serie de % por
    orden del producto. Devuelve las coordenadas listas para el <polyline> y el
    punto final, escaladas a un viewBox de 120x30. La forma se escala al propio
    rango del producto (min..max) porque interesa la tendencia, no el absoluto.
    """
    n = len(puntos)
    x0, x1, y_top, y_bot = 4.0, 116.0, 4.0, 26.0

    if n == 0:
        return {"n": 0, "puntos": "", "dot": None, "estilo": "solo",
                "stroke": _GRIS_SPARK, "dot_fill": _GRIS_SPARK}
    if n == 1:
        return {"n": 1, "puntos": "", "dot": [60.0, 15.0], "estilo": "solo",
                "stroke": _GRIS_SPARK, "dot_fill": _GRIS_SPARK}

    mn, mx = min(puntos), max(puntos)
    coords = []
    for i, p in enumerate(puntos):
        x = x0 + i * (x1 - x0) / (n - 1)
        if mx > mn:
            y = y_bot - (p - mn) / (mx - mn) * (y_bot - y_top)
        else:
            y = 15.0
        coords.append([round(x, 1), round(y, 1)])

    if n == 2:
        # Indicio: línea punteada gris (se estiliza en CSS/atributo).
        estilo, stroke, dot_fill = "punteado", _GRIS_SPARK, _GRIS_SPARK_DOT
    else:
        estilo = "linea"
        stroke = dot_fill = _NIVEL_HEX.get(nivel, _GRIS_SPARK)

    return {
        "n": n,
        "puntos": " ".join(f"{x},{y}" for x, y in coords),
        "dot": coords[-1],
        "estilo": estilo,
        "stroke": stroke,
        "dot_fill": dot_fill,
    }


def _direccion_tendencia(puntos):
    """
    Dirección de la tendencia del desperdicio de un producto:
      - <=1 punto -> 'na' ('— dato único')
      - 2 puntos  -> 'na' ('indicio (2 datos)')
      - 3+ puntos -> pendiente de una regresión lineal simple sobre TODOS los
        puntos; <0 'baja' (mejora), >0 'sube' (empeora), ~0 'estable'.
    """
    n = len(puntos)
    if n <= 1:
        return {"tipo": "na", "texto": "— dato único", "slope": None}
    if n == 2:
        return {"tipo": "na", "texto": "indicio (2 datos)", "slope": None}

    media_x = (n - 1) / 2.0
    media_y = sum(puntos) / n
    num = sum((i - media_x) * (p - media_y) for i, p in enumerate(puntos))
    den = sum((i - media_x) ** 2 for i in range(n))
    slope = (num / den) if den else 0.0

    if slope < -UMBRAL_PENDIENTE:
        return {"tipo": "baja", "texto": "Baja", "slope": round(slope, 4)}
    if slope > UMBRAL_PENDIENTE:
        return {"tipo": "sube", "texto": "Sube", "slope": round(slope, 4)}
    return {"tipo": "estable", "texto": "Estable", "slope": round(slope, 4)}


def _productos_resumen(ordenes):
    """
    Construye la lista de productos para el tab Resumen (barras horizontales y
    tabla de tendencia con sparklines), en memoria y sin nuevas consultas.

    Por producto (agrupado por COD): nombre, código, nº de órdenes, % ponderado
    global, nivel de color, ancho de barra relativo al máximo, sparkline SVG de
    la serie de % por orden (cronológica) y dirección por regresión lineal.
    Se ordena de mayor a menor % y se excluyen productos sin % calculable.
    """
    grupos = {}
    for orden in ordenes:
        cab = orden.get("cabecera") or {}
        cod = (str(cab.get("COD")).strip() if cab.get("COD") is not None else "") or "—"
        nom = str(cab.get("NOM")).strip() if cab.get("NOM") is not None else ""
        desp = orden.get("desperdicio") or {}
        g = grupos.get(cod)
        if g is None:
            g = grupos[cod] = {"cod": cod, "nom": nom, "n_ordenes": 0,
                               "num": 0.0, "den": 0.0, "puntos": []}
        if not g["nom"] and nom:
            g["nom"] = nom
        g["n_ordenes"] += 1
        num, den = desp.get("num_ponderado"), desp.get("den_ponderado")
        # Mismo criterio de ponderado que _agrupar_ponderado: num = kg de
        # desperdicio, den = base de la orden (consumo por nota; entregado para
        # 2A02). Solo entran las órdenes con base > 0.
        if num is not None and den is not None and float(den) > 0:
            g["num"] += float(num)
            g["den"] += float(den)
        pct = desp.get("porcentaje")
        if pct is not None:
            g["puntos"].append((_fecha_orden(cab), pct))

    productos = []
    for g in grupos.values():
        porcentaje = _porcentaje_ponderado(g["num"], g["den"])
        if porcentaje is None:
            continue  # sin % calculable -> no se grafica ni se lista
        # Serie de % por orden, en orden cronológico (fecha ascendente).
        serie = [p for _, p in sorted(g["puntos"], key=lambda t: t[0])]
        nivel = _nivel_desperdicio(porcentaje)
        productos.append({
            "cod": g["cod"],
            "nom": g["nom"] or "Sin nombre",
            "n_ordenes": g["n_ordenes"],
            "porcentaje": porcentaje,
            "nivel": nivel,
            "spark": _sparkline(serie, nivel),
            "direccion": _direccion_tendencia(serie),
        })

    productos.sort(key=lambda p: p["porcentaje"], reverse=True)
    # Ancho de barra proporcional al máximo % (el primero tras ordenar).
    maximo = productos[0]["porcentaje"] if productos else 0.0
    for p in productos:
        p["width"] = round(p["porcentaje"] / maximo * 100) if maximo > 0 else 0
    return productos


def _fecha_orden(cab):
    """Fecha de la orden como date ordenable; sentinela alto si no hay fecha."""
    f = cab.get("FECHA_I")
    if isinstance(f, datetime):
        return f.date()
    if hasattr(f, "year") and hasattr(f, "month") and hasattr(f, "day"):
        return f
    return date.max


def _resumen_ordenes(ordenes, analisis):
    """
    Construye los KPIs y los datos para los gráficos del tab "Resumen" a partir
    de los resultados ya obtenidos en memoria (NO lanza consultas nuevas a SQL
    Server). Reutiliza la agregación por producto de `analisis`.

    Devuelve un dict apto para serializar a JSON y consumir desde Chart.js:
    - total: nº de órdenes de la consulta.
    - max / min: {op, porcentaje} de la orden con mayor / menor % (None si
      ninguna orden tiene % calculado).
    - promedio_ponderado: suma(valor_nota) / suma(CANT1) * 100 del conjunto.
    - grafico_producto / grafico_tiempo: {labels, data} listos para graficar.
    """
    total = len(ordenes)

    # KPIs máximo/mínimo (sobre órdenes con % calculado) y ponderado general.
    max_kpi = None
    min_kpi = None
    num_total = 0.0
    den_total = 0.0
    for orden in ordenes:
        desp = orden.get("desperdicio") or {}
        pct = desp.get("porcentaje")
        num = desp.get("num_ponderado")
        den = desp.get("den_ponderado")
        # Mismo criterio de ponderado: num = kg de desperdicio, den = base de la
        # orden (consumo por nota; entregado para 2A02). Solo con base > 0.
        if num is not None and den is not None and float(den) > 0:
            num_total += float(num)
            den_total += float(den)
        if pct is not None:
            if max_kpi is None or pct > max_kpi["porcentaje"]:
                max_kpi = {"op": orden.get("op"), "porcentaje": pct}
            if min_kpi is None or pct < min_kpi["porcentaje"]:
                min_kpi = {"op": orden.get("op"), "porcentaje": pct}

    promedio = _porcentaje_ponderado(num_total, den_total)

    # Gráfico por producto: reutiliza la agregación (ya ordenada desc) del tab
    # Análisis; solo se grafican los grupos con % calculable.
    prod = [f for f in analisis["por_producto"] if f["porcentaje"] is not None]
    grafico_producto = {
        "labels": [f["etiqueta"] for f in prod],
        "data": [f["porcentaje"] for f in prod],
    }

    # Gráfico temporal: por día si el rango es corto, por mes si es amplio.
    fechas = _fechas_validas(ordenes)
    agrupacion = "mes"
    if fechas and (max(fechas) - min(fechas)).days <= UMBRAL_DIAS_AGRUPAR_MES:
        agrupacion = "día"
    clave = _clave_mes if agrupacion == "mes" else _clave_dia
    temporal = _agrupar_ponderado(ordenes, clave)
    temporal.sort(key=lambda f: f["id"])  # cronológico; "Sin fecha" al final
    temporal = [f for f in temporal if f["porcentaje"] is not None]
    grafico_tiempo = {
        "labels": [f["etiqueta"] for f in temporal],
        "data": [f["porcentaje"] for f in temporal],
        "agrupacion": agrupacion,
    }

    # Tendencia por producto (líneas): reutiliza la agregación en memoria y el
    # orden de productos del gráfico de barras para colores coherentes.
    tendencia = _tendencia_por_grupo(ordenes, grafico_producto["labels"], agrupacion)

    # Productos para el tab Resumen v2 (barras horizontales + tabla sparklines).
    productos = _productos_resumen(ordenes)

    return {
        "total": total,
        "max": max_kpi,
        "min": min_kpi,
        "promedio_ponderado": promedio,
        "grafico_producto": grafico_producto,
        "grafico_tiempo": grafico_tiempo,
        "tendencia": tendencia,
        "productos": productos,
    }


def _analizar_ordenes(ordenes):
    """
    Construye el análisis agregado sobre los resultados ya obtenidos en memoria
    (NO lanza consultas nuevas a SQL Server). Devuelve un dict con tres bloques:

    - por_producto: agrupado por COD/NOM, ordenado de mayor a menor % ponderado.
    - por_mes: agrupado por año-mes de FECHA_I, ordenado cronológicamente.
    - por_cliente: pendiente hasta que la consulta de OP incorpore el cliente;
      la estructura ya queda lista para poblarla sin rehacer el tab.
    """
    por_producto = _agrupar_ponderado(ordenes, _clave_producto)
    # Mayor desperdicio primero; None (sin base) al final.
    por_producto.sort(
        key=lambda f: (f["porcentaje"] is None, -(f["porcentaje"] or 0.0))
    )

    por_mes = _agrupar_ponderado(ordenes, _clave_mes)
    por_mes.sort(key=lambda f: f["id"])  # cronológico; "Sin fecha" (9999-99) al final

    # --- Por cliente (Opción B: grupo conjunto) ---
    # Se agrupa vía la relación CIFRADA producto->cliente(s). Si la clave falta
    # o el archivo no es descifrable, se degrada con un mensaje claro sin
    # romper el resto del análisis (producto/mes siguen funcionando).
    try:
        mapa_clientes = cargar_clientes()
        clave_cliente = _hacer_clave_cliente(mapa_clientes)
        # Se anota el id de grupo en cada orden para que el acordeón del tab
        # empareje sus miembros sin rehacer el mapeo en la plantilla.
        for orden in ordenes:
            orden["grupo_cliente"] = clave_cliente(orden)[0]
        filas_cliente = _agrupar_ponderado(ordenes, clave_cliente)
        # Mayor desperdicio primero; "Sin cliente asignado" y % None al final.
        filas_cliente.sort(
            key=lambda f: (
                f["id"] == "Sin cliente asignado",
                f["porcentaje"] is None,
                -(f["porcentaje"] or 0.0),
            )
        )
        por_cliente = {"disponible": True, "mensaje": None, "filas": filas_cliente}
    except ClientesError as exc:
        por_cliente = {
            "disponible": False,
            "mensaje": f"Análisis por cliente no disponible: {exc}",
            "filas": [],
        }

    return {
        "por_producto": por_producto,
        "por_mes": por_mes,
        "por_cliente": por_cliente,
    }


def _ordenes_data_json(ordenes):
    """
    Dataset por orden para el motor de agregación en cliente (buscador que
    re-renderiza Resultado/Análisis/Resumen sin reconsultar). Se llama DESPUÉS
    de _analizar_ordenes (que anota `grupo_cliente` en cada orden) para poder
    incluirlo. NO recalcula nada: solo lee valores que el backend YA calculó
    (num_ponderado/den_ponderado/porcentaje de la regla de cada orden, el
    mismo grupo_cliente Opción B de Análisis).

    `mes` usa FECHA_I (igual que _clave_mes; FECHA_T es solo para resolver la
    tabla Kardex, un asunto distinto). `fecha` (ISO YYYY-MM-DD o None) es un
    campo EXTRA no pedido explícitamente, necesario para que el sparkline de
    "Tendencia por producto" ordene los puntos cronológicamente igual que
    _fecha_orden — con "mes" solo, dos órdenes del mismo mes quedarían en
    orden arbitrario y el sparkline no coincidiría con el de Python.
    `suma_cant1`/`desperdicio_kg`/`unidad` también se agregan (ya calculados,
    cero costo) para que las sub-tablas de detalle del acordeón (clic en una
    fila de Análisis) puedan reconstruirse completas en JS, igual que
    ui.detalle_grupo hoy en Jinja.

    IMPORTANTE: `suma_cant1`/`den_ponderado` de la regla por NOTA vienen tal
    cual de pyodbc (decimal.Decimal, ver _sumar_cant1_2a), no como float — a
    diferencia de las reglas 2A02/kardex que ya normalizan. `tojson`/`json`
    NO sabe serializar Decimal, así que aquí se fuerza float() a los 4 campos
    numéricos (ninguno debe llegar nunca como Decimal al HTML).
    """
    def _num(v):
        return float(v) if v is not None else None

    datos = []
    for orden in ordenes:
        cab = orden.get("cabecera") or {}
        desp = orden.get("desperdicio") or {}
        fecha_i = cab.get("FECHA_I")
        # Mismo criterio que _clave_mes: pyodbc puede devolver datetime.datetime
        # O datetime.date "pelado" (sin hora) según el tipo de columna en SQL
        # Server; hay que aceptar ambos, no solo datetime.
        if isinstance(fecha_i, datetime) or (
            hasattr(fecha_i, "year") and hasattr(fecha_i, "month") and hasattr(fecha_i, "day")
        ):
            mes = f"{fecha_i.year:04d}-{fecha_i.month:02d}"
            fecha_iso = f"{fecha_i.year:04d}-{fecha_i.month:02d}-{fecha_i.day:02d}"
        else:
            mes = "9999-99"
            fecha_iso = None
        datos.append({
            "op": orden.get("op"),
            "cod": cab.get("COD"),
            "nombre": cab.get("NOM"),
            "mes": mes,
            "fecha": fecha_iso,
            "grupo_cliente": orden.get("grupo_cliente") or "Sin cliente asignado",
            "num_ponderado": _num(desp.get("num_ponderado")),
            "den_ponderado": _num(desp.get("den_ponderado")),
            "tiene_calculo": desp.get("porcentaje") is not None,
            "pct": desp.get("porcentaje"),
            "suma_cant1": _num(desp.get("suma_cant1")),
            "desperdicio_kg": _num(desp.get("desperdicio_kg")),
            "unidad": desp.get("unidad"),
        })
    return datos


@app.route("/ordenes-produccion/consultar", methods=["POST"])
@login_required
def ordenes_produccion_consultar():
    """
    Ejecuta la consulta y renderiza los resultados agrupados por orden.
    """
    try:
        ordenes = _consultar_ordenes(request.form)
    except ValueError as exc:
        return render_template(
            "ordenes_resultado.html", error=str(exc), ordenes=[]
        ), 400
    except Exception as exc:
        return render_template(
            "ordenes_resultado.html",
            error=str(exc),
            traceback=traceback.format_exc(),
            ordenes=[],
        ), 500

    analisis = _analizar_ordenes(ordenes)
    return render_template(
        "ordenes_resultado.html",
        error=None,
        ordenes=ordenes,
        analisis=analisis,
        resumen=_resumen_ordenes(ordenes, analisis),
        filtro=_extraer_filtro(request.form),
        ordenes_data=_ordenes_data_json(ordenes),
    ), 200


def _generar_excel(ordenes):
    """
    Construye un .xlsx en memoria: una sola tabla con una fila por OP y las
    columnas OP, Fecha, Metros, Kg compuesto, Kg desperdicio, Unidad y
    % desperdicio. La columna Unidad aclara en qué unidad están las dos
    anteriores: 'KG' para las reglas por nota/diferencia, o la unidad real del
    movimiento (p. ej. 'UN') para la regla farmacéutica (Kardex).
    Si una OP tiene alerta, el % y los Kg desperdicio quedan con el texto de la
    alerta / vacíos. Devuelve un BytesIO listo para send_file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Producción"

    header_font = Font(bold=True, color="334155")
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    left = Alignment(horizontal="left", vertical="center")

    encabezados = [
        "OP",
        "Código",
        "Nombre",
        "Fecha",
        "Metros",
        "Kg compuesto",
        "Kg desperdicio",
        "Unidad",
        "% desperdicio",
    ]
    for i, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=i, value=texto)
        c.font = header_font
        c.fill = header_fill
        c.alignment = left

    fila = 2
    for orden in ordenes:
        cab = orden.get("cabecera") or {}
        d = orden.get("desperdicio", {})
        alerta = d.get("alerta")

        # Kg de desperdicio method-agnostic: la nota (regla general) o la
        # diferencia consumido-entregado (regla 2A02). None si hay alerta.
        if alerta:
            kg_desperdicio = d.get("desperdicio_kg")  # None si no fue calculable
            porcentaje = alerta
        else:
            kg_desperdicio = d.get("desperdicio_kg")
            porcentaje = d.get("porcentaje")

        # La fecha viene como datetime; nos quedamos solo con la parte de fecha
        # para que Excel no muestre hora/minutos/segundos.
        fecha_i = cab.get("FECHA_I")
        if isinstance(fecha_i, datetime):
            fecha_i = fecha_i.date()

        valores = [
            orden.get("op"),
            cab.get("COD"),        # Código
            cab.get("NOM"),        # Nombre
            fecha_i,               # Fecha (sin hora)
            cab.get("CANTE"),      # Metros (para 2A02 = kg entregado)
            d.get("suma_cant1"),   # Consumo (2A0/2A62; 2A02=mezcla 2A01; Kardex=Σ CANT)
            kg_desperdicio,        # Desperdicio (nota, diferencia o Kardex)
            d.get("unidad"),       # Unidad del consumo/desperdicio (KG, UN, ...)
            porcentaje,            # % o texto de alerta
        ]
        for i, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=i, value=valor)
            # Columna 4 = Fecha: formato de fecha sin componente de hora.
            if i == 4 and valor is not None:
                celda.number_format = "yyyy-mm-dd"
        fila += 1

    # Anchos de columna para legibilidad
    anchos = [12, 14, 30, 14, 12, 16, 16, 10, 26]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


@app.route("/ordenes-produccion/exportar", methods=["POST"])
@login_required
def ordenes_produccion_exportar():
    """
    Reejecuta la consulta con el mismo filtro y devuelve un archivo .xlsx.
    """
    try:
        ordenes = _consultar_ordenes(request.form)
    except ValueError as exc:
        return f"Filtro inválido: {exc}", 400
    except Exception as exc:
        return f"Error al consultar la base de datos: {exc}", 500

    excel = _generar_excel(ordenes)
    nombre = f"ordenes_produccion_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return send_file(
        excel,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _criterios_reporte(filtro):
    """Texto legible de los criterios de la consulta para la portada del PDF."""
    if not filtro:
        return "Todos los resultados"
    if filtro.get("modo") == "orden":
        ops = ", ".join(str(o) for o in filtro.get("ordenes", []))
        return f"Órdenes (OP): {ops}"
    return (f"Producto {filtro.get('cod', '')}% · "
            f"Fechas {filtro.get('fecha_desde', '')} a {filtro.get('fecha_hasta', '')}")


def _generar_pdf(ordenes, analisis, resumen, filtro):
    """
    Construye el "Reporte de desperdicio" en PDF (en servidor) reutilizando los
    datos ya calculados por _consultar_ordenes / _analizar_ordenes /
    _resumen_ordenes. Texto real seleccionable, paginado, con la paleta índigo
    del aplicativo. Devuelve un BytesIO. Importa ReportLab de forma perezosa
    para que su ausencia solo afecte a esta ruta, no al arranque de la app.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfgen import canvas as _canvas
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, LongTable)

    # --- Paleta (design tokens del aplicativo) ---
    PRIMARIO = colors.HexColor("#4F46E5")
    PRIMARIO_OSC = colors.HexColor("#4338CA")
    PRIMARIO_SUAVE = colors.HexColor("#EEF0FE")
    TEXTO = colors.HexColor("#141519")
    TEXTO2 = colors.HexColor("#5C6270")
    TEXTO3 = colors.HexColor("#9296A1")
    BORDE = colors.HexColor("#EAEBEF")
    FONDO = colors.HexColor("#FBFBFC")
    VERDE = colors.HexColor("#0E9F6E")
    AMBER = colors.HexColor("#E8A317")
    NARANJA = colors.HexColor("#F07A1E")
    # Colores de texto del % (variantes -txt, legibles sobre fondo claro).
    HEX_PCT = {"verde": "#0E9F6E", "amber": "#A67908", "naranja": "#D9631A"}
    BAR_COL = {"verde": VERDE, "amber": AMBER, "naranja": NARANJA}

    fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M")
    ancho_util = A4[0] - 36 * mm  # márgenes de 18mm por lado

    def esc(s):
        return (str(s) if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def hex_pct(pct):
        return HEX_PCT.get(_nivel_desperdicio(pct), "#5C6270")

    def fmt_num(x):
        if x is None:
            return "—"
        try:
            f = float(x)
        except (TypeError, ValueError):
            return esc(x)
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
        return f"{f:.2f}"

    def fmt_fecha(f):
        if isinstance(f, datetime) or hasattr(f, "strftime"):
            try:
                return f.strftime("%Y-%m-%d")
            except (ValueError, AttributeError):
                return "—"
        return "—"

    # --- Estilos de párrafo ---
    S = {
        "title": ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=20,
                                textColor=PRIMARIO_OSC, leading=23, spaceAfter=3),
        "sub": ParagraphStyle("sub", fontName="Helvetica", fontSize=9.5, textColor=TEXTO2, leading=13),
        "h2": ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=13, textColor=TEXTO,
                            leading=16, spaceBefore=16, spaceAfter=8),
        "cell": ParagraphStyle("cell", fontName="Helvetica", fontSize=8.5, textColor=TEXTO, leading=11),
        "cellr": ParagraphStyle("cellr", fontName="Helvetica", fontSize=8.5, textColor=TEXTO, leading=11, alignment=2),
        "cellc": ParagraphStyle("cellc", fontName="Helvetica", fontSize=8.5, textColor=TEXTO, leading=11, alignment=1),
        "th": ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=7.5, textColor=TEXTO3, leading=10),
        "thr": ParagraphStyle("thr", fontName="Helvetica-Bold", fontSize=7.5, textColor=TEXTO3, leading=10, alignment=2),
        "thc": ParagraphStyle("thc", fontName="Helvetica-Bold", fontSize=7.5, textColor=TEXTO3, leading=10, alignment=1),
        "klab": ParagraphStyle("klab", fontName="Helvetica", fontSize=8, textColor=TEXTO2, leading=10),
        "kval": ParagraphStyle("kval", fontName="Helvetica-Bold", fontSize=19, textColor=TEXTO, leading=21, spaceBefore=3),
        "ksub": ParagraphStyle("ksub", fontName="Helvetica", fontSize=7.5, textColor=TEXTO3, leading=10, spaceBefore=2),
        "empty": ParagraphStyle("empty", fontName="Helvetica", fontSize=12, textColor=TEXTO2, alignment=1, leading=18),
    }

    # --- Canvas con encabezado/pie y numeración "Página X de Y" ---
    class _CanvasReporte(_canvas.Canvas):
        def __init__(self, *a, **k):
            super().__init__(*a, **k)
            self._paginas = []

        def showPage(self):
            self._paginas.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._paginas)
            for i, estado in enumerate(self._paginas, start=1):
                self.__dict__.update(estado)
                self._encabezado_pie(i, total)
                super().showPage()
            super().save()

        def _encabezado_pie(self, num, total):
            w, h = A4
            lm, rm = 18 * mm, 18 * mm
            # Encabezado: marca PLASTITEC + módulo
            self.setFillColor(PRIMARIO)
            self.setFont("Helvetica-Bold", 11)
            self.drawString(lm, h - 15 * mm, "PLASTITEC")
            x2 = lm + self.stringWidth("PLASTITEC", "Helvetica-Bold", 11) + 5
            self.setFillColor(TEXTO3)
            self.setFont("Helvetica", 9)
            self.drawString(x2, h - 15 * mm, "— Cálculo de desperdicio")
            self.setStrokeColor(BORDE)
            self.setLineWidth(0.7)
            self.line(lm, h - 17 * mm, w - rm, h - 17 * mm)
            # Pie: fecha de generación + numeración
            self.line(lm, 15 * mm, w - rm, 15 * mm)
            self.setFillColor(TEXTO3)
            self.setFont("Helvetica", 8)
            self.drawString(lm, 10.5 * mm, f"Generado: {fecha_gen}")
            self.drawRightString(w - rm, 10.5 * mm, f"Página {num} de {total}")

    flow = []
    flow.append(Paragraph("Reporte de desperdicio", S["title"]))
    flow.append(Paragraph(esc(_criterios_reporte(filtro)), S["sub"]))
    flow.append(Paragraph(f"Fecha de generación: {fecha_gen}", S["sub"]))
    flow.append(Spacer(1, 4 * mm))

    # --- Caso límite: sin resultados ---
    if not ordenes:
        flow.append(Spacer(1, 40 * mm))
        flow.append(Paragraph(
            "Sin datos para el reporte.<br/>"
            "La consulta no devolvió órdenes con el filtro indicado.", S["empty"]))
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                                topMargin=24 * mm, bottomMargin=20 * mm,
                                title="Reporte de desperdicio")
        doc.build(flow, canvasmaker=_CanvasReporte)
        buf.seek(0)
        return buf

    productos = resumen.get("productos", [])

    # --- 2. Resumen ejecutivo (KPIs) ---
    flow.append(Paragraph("Resumen ejecutivo", S["h2"]))

    def kpi_cell(label, valor, color_hex, sub):
        return [
            Paragraph(esc(label), S["klab"]),
            Paragraph(f'<font color="{color_hex}">{esc(valor)}</font>', S["kval"]),
            Paragraph(esc(sub), S["ksub"]),
        ]

    maxk, mink = resumen.get("max"), resumen.get("min")
    prom = resumen.get("promedio_ponderado")
    kpis = [[
        kpi_cell("Órdenes encontradas", resumen.get("total", 0), "#141519",
                 f"{len(analisis.get('por_producto', []))} producto(s)"),
        kpi_cell("Desperdicio más alto", f"{maxk['porcentaje']}%" if maxk else "—",
                 hex_pct(maxk["porcentaje"]) if maxk else "#5C6270",
                 f"OP {maxk['op']}" if maxk else "sin % calculado"),
        kpi_cell("Desperdicio más bajo", f"{mink['porcentaje']}%" if mink else "—",
                 hex_pct(mink["porcentaje"]) if mink else "#5C6270",
                 f"OP {mink['op']}" if mink else "sin % calculado"),
        kpi_cell("Promedio ponderado", f"{prom}%" if prom is not None else "—",
                 hex_pct(prom) if prom is not None else "#5C6270", "sobre volumen total"),
    ]]
    t_kpi = Table(kpis, colWidths=[ancho_util / 4.0] * 4)
    t_kpi.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, BORDE),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 11), ("BOTTOMPADDING", (0, 0), (-1, -1), 11),
    ]))
    flow.append(t_kpi)
    flow.append(Spacer(1, 3 * mm))
    flow.append(Paragraph(
        "El % ponderado se calcula solo sobre las órdenes con nota de desperdicio; "
        "el consumo de órdenes sin nota no se incluye en el denominador.", S["ksub"]))

    # --- 3. Ranking de desperdicio por producto (tabla + barras) ---
    flow.append(Paragraph("Ranking de desperdicio por producto", S["h2"]))
    if productos:
        maximo = productos[0]["porcentaje"] or 0
        datos = [[Paragraph("PRODUCTO", S["th"]), Paragraph("ÓRDENES", S["thc"]),
                  Paragraph("NIVEL", S["th"]), Paragraph("% DESP.", S["thr"])]]
        for p in productos:
            barw = (p["porcentaje"] / maximo * 42 * mm) if maximo else 0
            barw = max(barw, 1.5)
            barra = Table([[""]], colWidths=[barw], rowHeights=[5 * mm])
            barra.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), BAR_COL.get(p["nivel"], TEXTO3)),
                ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            datos.append([
                Paragraph(f'{esc(p["nom"])}<br/><font color="#9296A1" size="7">{esc(p["cod"])}</font>', S["cell"]),
                Paragraph(str(p["n_ordenes"]), S["cellc"]),
                barra,
                Paragraph(f'<b><font color="{hex_pct(p["porcentaje"])}">{p["porcentaje"]}%</font></b>', S["cellr"]),
            ])
        t_rank = Table(datos, colWidths=[70 * mm, 24 * mm, 50 * mm, 30 * mm], repeatRows=1)
        t_rank.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARIO_SUAVE),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 1), (2, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        flow.append(t_rank)
        flow.append(Spacer(1, 3 * mm))
        flow.append(Paragraph(
            '<font color="#0E9F6E">■</font> Aceptable (&lt;10%)&nbsp;&nbsp;'
            '<font color="#E8A317">■</font> Atención (10–15%)&nbsp;&nbsp;'
            '<font color="#F07A1E">■</font> Alto (&gt;15%)', S["ksub"]))
    else:
        flow.append(Paragraph("Sin productos con desperdicio calculable.", S["sub"]))

    # --- 4. Tendencia por producto (dirección + % actual) ---
    flow.append(Paragraph("Tendencia por producto", S["h2"]))
    if productos:
        DIR = {"baja": ("↓ Baja", "#0E9F6E"), "sube": ("↑ Sube", "#D9631A"),
               "estable": ("Estable", "#9296A1")}
        datos = [[Paragraph("PRODUCTO", S["th"]), Paragraph("% ACTUAL", S["thr"]),
                  Paragraph("DIRECCIÓN", S["thr"])]]
        for p in productos:
            dtipo = p["direccion"]["tipo"]
            if dtipo in DIR:
                dtxt, dcol = DIR[dtipo]
            else:
                dtxt, dcol = esc(p["direccion"]["texto"]), "#9296A1"
            datos.append([
                Paragraph(f'{esc(p["nom"])}<br/><font color="#9296A1" size="7">{esc(p["cod"])} · {p["n_ordenes"]} orden(es)</font>', S["cell"]),
                Paragraph(f'<b><font color="{hex_pct(p["porcentaje"])}">{p["porcentaje"]}%</font></b>', S["cellr"]),
                Paragraph(f'<font color="{dcol}">{dtxt}</font>', S["cellr"]),
            ])
        t_tend = Table(datos, colWidths=[104 * mm, 30 * mm, 40 * mm], repeatRows=1)
        t_tend.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARIO_SUAVE),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, BORDE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        flow.append(t_tend)
        flow.append(Spacer(1, 2 * mm))
        flow.append(Paragraph(
            "Baja = el desperdicio mejora · Sube = empeora. "
            "La dirección solo se calcula con 3+ órdenes.", S["ksub"]))
    else:
        flow.append(Paragraph("Sin productos con desperdicio calculable.", S["sub"]))

    # --- 5. Detalle de órdenes (tabla paginada) ---
    flow.append(Paragraph("Detalle de órdenes", S["h2"]))
    datos = [[Paragraph("OP", S["th"]), Paragraph("CÓDIGO", S["th"]), Paragraph("NOMBRE", S["th"]),
              Paragraph("FECHA", S["th"]), Paragraph("CONSUMO", S["thr"]),
              Paragraph("DESPERD.", S["thr"]), Paragraph("%", S["thr"])]]
    for orden in ordenes:
        cab = orden.get("cabecera") or {}
        d = orden.get("desperdicio") or {}
        pct = d.get("porcentaje")
        if pct is not None:
            pct_par = Paragraph(f'<b><font color="{hex_pct(pct)}">{pct}%</font></b>', S["cellr"])
        else:
            pct_par = Paragraph(f'<font color="#9296A1" size="7">{esc(d.get("alerta") or "—")}</font>', S["cellr"])
        # Farmacéutico (Kardex): la unidad puede NO ser kg (p. ej. UN), se
        # aclara junto al número. Las demás reglas siguen mostrando el número
        # sin sufijo, como antes (siempre son kg).
        sufijo_unidad = f" {d['unidad']}" if d.get("metodo") == "kardex" and d.get("unidad") else ""

        def fmt_con_unidad(x):
            txt = fmt_num(x)
            return txt + sufijo_unidad if (txt != "—" and sufijo_unidad) else txt

        datos.append([
            Paragraph(esc(orden.get("op")), S["cell"]),
            Paragraph(esc(cab.get("COD")), S["cell"]),
            Paragraph(esc(cab.get("NOM")), S["cell"]),
            Paragraph(fmt_fecha(cab.get("FECHA_I")), S["cell"]),
            Paragraph(esc(fmt_con_unidad(d.get("suma_cant1"))), S["cellr"]),
            Paragraph(esc(fmt_con_unidad(d.get("desperdicio_kg"))), S["cellr"]),
            pct_par,
        ])
    t_det = LongTable(datos, colWidths=[16 * mm, 24 * mm, 54 * mm, 22 * mm, 24 * mm, 18 * mm, 16 * mm], repeatRows=1)
    t_det.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARIO_SUAVE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, FONDO]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, BORDE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    flow.append(t_det)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=24 * mm, bottomMargin=20 * mm,
                            title="Reporte de desperdicio", author="PLASTITEC")
    doc.build(flow, canvasmaker=_CanvasReporte)
    buf.seek(0)
    return buf


@app.route("/ordenes-produccion/reporte-pdf", methods=["POST"])
@login_required
def ordenes_produccion_reporte_pdf():
    """
    Regenera los datos con el MISMO filtro que la consulta (reutilizando las
    funciones de consulta y agregación) y devuelve el reporte en PDF.
    """
    try:
        ordenes = _consultar_ordenes(request.form)
    except ValueError as exc:
        return f"Filtro inválido: {exc}", 400
    except Exception as exc:
        return f"Error al consultar la base de datos: {exc}", 500

    analisis = _analizar_ordenes(ordenes)
    resumen = _resumen_ordenes(ordenes, analisis)
    filtro = _extraer_filtro(request.form)

    try:
        pdf = _generar_pdf(ordenes, analisis, resumen, filtro)
    except ImportError:
        return ("La generación de PDF requiere la librería 'reportlab'. "
                "Instálala con: pip install reportlab==5.0.0"), 500

    nombre = f"reporte_desperdicio_{datetime.now():%Y%m%d_%H%M}.pdf"
    return send_file(pdf, as_attachment=True, download_name=nombre, mimetype="application/pdf")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
